import openpyxl

# 読み込むファイルのデータを確認する
wb = openpyxl.load_workbook('sample.xlsx')

ws = wb.active

# ファイルの内容を確認する（行単位で値を取得）
for row in ws.iter_rows(min_row=1, min_col=1):
    for cell in row:
        print(cell.value)

# ファイルの内容を確認する（列単位で値を取得）
for col in ws.iter_cols(min_row=1, min_col=1):
    for cell in col:
        print(cell.value)

### 定刻よりも早く出発（到着）していれば◯、していなければ✕を表す列を追加する ###

# 各列(D, F)の出発時間を取得する
STD = []
for col in ws.iter_cols(min_row=2, min_col=4, max_col=4):
    for cell in col:
        STD.append(cell.value)

ATD = []
for col in ws.iter_cols(min_row=2, min_col=6, max_col=6):
    for cell in col:
        ATD.append(cell.value)

# 値が格納されているか確認する
for i, k in zip(STD, ATD):
    print(i, k)

# 時間を判定する
DEP = []
for i, k in zip(STD, ATD):
    if i >= k:
        DEP.append('◯')
    else:
        DEP.append('✕')

print(len(DEP))

# 判定結果を新しい列に書き込んで保存する
ws['I1'] = '定時出発'
for col in ws.iter_cols(min_row=2, min_col=9, max_col=9):
    for cell, i in zip(col, DEP):
        cell.value = i

wb.save('sample.xlsx')

# 各列(E, G)の到着時間を取得する
STA = []
for col in ws.iter_cols(min_row=2, min_col=5, max_col=5):
    for cell in col:
        STA.append(cell.value)

ATA = []
for col in ws.iter_cols(min_row=2, min_col=7, max_col=7):
    for cell in col:
        ATA.append(cell.value)

# 値が格納されているか確認する
for i, k in zip(STA, ATA):
    print(i, k)

# 時間を判定する
ARR = []
for i, k in zip(STA, ATA):
    if i >= k:
        ARR.append('◯')
    else:
        ARR.append('✕')

# 判定結果を新しい列に書き込む
ws['J1'] = '定時到着'
for col in ws.iter_cols(min_row=2, min_col=10, max_col=10):
    for cell, i in zip(col, ARR):
        cell.value = i

wb.save('sample.xlsx')

### 乗車率を表す列を作成する ###

# 最大席数を45と仮定して、各便の乗車率を計算する（乗客数÷45）
ws['K1'] = '乗車率'

LOAD_FACTOR = []
for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
    for cell in col:
        lf = cell.value / 45
        LOAD_FACTOR.append('{:.0%}'.format(lf))

print(LOAD_FACTOR)

for col in ws.iter_cols(min_row=2, min_col=11,
                        max_row=len(LOAD_FACTOR)+1, max_col=11):
    for cell, i in zip(col, LOAD_FACTOR):
        cell.value = i

wb.save('sample.xlsx')

# max_rowの返りを確認する
for col in ws.iter_cols(min_row=2, min_col=20,
                        max_row=10, max_col=20):
    for cell in col:
        print(cell)

# グラフを反映するシートを追加する
ws2 = wb.create_sheet('chart')

# 追加されているか確認する
print(wb.sheetnames)

# グラフのためのデータを作成する
# 区間ごとの運行数を作成する
route_num = {}
for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
    for cell in col:
        if cell.value in route_num:
            route_num[cell.value] += 1
        else:
            route_num[cell.value] = 1

# 値が格納されているか確認する
for i, j in route_num.items():
    print(i, j)

# chartシートにデータを反映する
route = [i for i in route_num.keys()]
route.insert(0, '区間')

num = [i for i in route_num.values()]
num.insert(0, '運行数')

for col in ws2.iter_cols(min_row=1, min_col=1, max_row=len(route), max_col=1):
    for cell, r in zip(col, route):
        cell.value = r

for col in ws2.iter_cols(min_row=1, min_col=2, max_row=len(num), max_col=2):
    for cell, n in zip(col, num):
        cell.value = n
wb.save('sample.xlsx')

# 定時出発率、定時到着率、平均乗車率をchartシートに反映する
ws2['D1'] = '定時出発率'
dep_result = []
for col in ws.iter_cols(min_row=2, min_col=9, max_col=9):
    for cell in col:
        dep_result.append(cell.value)

dep_rate = dep_result.count('◯') / len(dep_result)
ws2['D2'] = '{:.0%}'.format(dep_rate)

ws2['E1'] = '定時到着率'
arr_result = []
for col in ws.iter_cols(min_row=2, min_col=10, max_col=10):
    for cell in col:
        arr_result.append(cell.value)

arr_rate = arr_result.count('◯') / len(arr_result)
ws2['E2'] = '{:.0%}'.format(arr_rate)

ws2['F1'] = '平均乗車率'
lf_result = 0
for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
    for cell in col:
        lf_result += cell.value / 45
lf_result = lf_result / len(arr_result)
ws2['F2'] = '{:.0%}'.format(lf_result)

wb.save('sample.xlsx')

### グラフを挿入する ###

# 必要な機能をimportする
from openpyxl.chart import BarChart, Reference, Series
    
# 棒グラフを初期化する
chart = BarChart()

# 描画するための値を作成
plot_value = Reference(ws2, min_row=1, min_col=2, max_row=5, max_col=2)
chart.add_data(plot_value, titles_from_data=True)

# X軸のラベルを作成
x = Reference(ws2, min_row=2, min_col=1, max_row=5, max_col=1)
chart.set_categories(x)
    
ws2.add_chart(chart, 'E5')

wb.save('sample.xlsx')

# 3Dの棒グラフを作成する
from openpyxl.chart import BarChart3D
chart3d = BarChart3D()

# 描画するための値を作成
plot_value = Reference(ws2, min_row=1, min_col=2, max_row=5, max_col=2)
chart3d.add_data(plot_value, titles_from_data=True)
chart3d.title = '3D Version'

# X軸のラベルを作成
x = Reference(ws2, min_row=2, min_col=1, max_row=5, max_col=1)
chart3d.set_categories(x)
    
ws2.add_chart(chart3d, 'N5')
wb.save('sample.xlsx')

