import openpyxl

### openpyxlからExcelファイルを作成する ###

# 最初にworkbookオブジェクトを生成する
wb = openpyxl.Workbook()

# 次にworkbookオブジェクトからworksheetオブジェクトを作成する
ws = wb.worksheets

print(wb.sheetnames)

wb = openpyxl.Workbook()

# workbookオブジェクト.activeでworksheetを取得
ws = wb.active

print(wb.sheetnames)

#　生成したworksheetオブジェクトに値を入力する
ws['A1'] = 'RX-105'

#　番地指定で値を入力する
ws.cell(row=2, column=1, value='RX-104FF')

# 保存
wb.save('test1.xlsx')

# 複数のセルに値を入力する
data = ['fruits', 'apple', 'orange', 'peach', 'melon', 'grape']

for row, i in zip(ws.iter_rows(min_row=1, min_col=2,
                               max_row=len(data), max_col=2), data):
    for cell in row:
        cell.value = i
        print(cell.value)

# 保存
wb.save('test1.xlsx')

# iter_rows()メソッドの返しを確認
for row in ws.iter_rows(min_row=1, min_col=3,
                               max_row=5, max_col=4):
    for cell in row:
        print(cell)

# 既存のExcelファイルを読み込む
wb2 = openpyxl.load_workbook('test2.xlsx')
ws2 = wb2.worksheets[0]

# 値を読み込む
for row in ws2.iter_rows(min_row=1, min_col=1, max_col=1):
    for cell in row:
        print(cell, "-", cell.value)

# 行を指定して特定の範囲の値を取得する
for row in ws2.iter_rows(min_row=3, min_col=1, max_row=8, max_col=1):
    for cell in row:
        print(cell, "-", cell.value)

# 取得したセルのデータを別のセルに書き込む
data = []

for cell in list(ws2.columns)[0]:
    data.append(cell.value)

print(data)

for col in ws2.iter_cols(min_row=1, min_col=2,
                         max_row=len(data), max_col=2):
    for cell, i in zip(col, data):
        cell.value = i

wb2.save('test3.xlsx')

# iter_cols()メソッドの返しを確認
for col in ws.iter_cols(min_row=1, min_col=3,
                               max_row=5, max_col=4):
    for cell in col:
        print(cell)
        